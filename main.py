import concurrent.futures
import pandas as pd
import portalocker
import os
import random
import threading
import time
from io import BytesIO
from functools import wraps
from openpyxl import load_workbook, Workbook
from pathlib import Path


DATA_INTERVAL = (0, 666)
SLEEP_INTERVAL = (0, 20)
THREADS = 3


def wait_for_file_lock(**lock_kwargs):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            wait = True
            file_path, *rest = args
            is_new_file = not os.path.isfile(file_path)

            if "mode" in lock_kwargs and "x" in lock_kwargs["mode"]:
                if is_new_file:
                    Path(file_path).touch()

                if "a" in lock_kwargs["mode"] or "r" in lock_kwargs["mode"]:
                    lock_kwargs["mode"] = lock_kwargs["mode"].replace("x", "")

            while wait:
                try:
                    with portalocker.Lock(file_path, **lock_kwargs) as fl:
                        wait = False
                        result = func(
                            *args,
                            **{"is_new_file": is_new_file, "locked_file": fl, **kwargs},
                        )
                except portalocker.LockException as ex:
                    print(f"File {file_path} is LOCKED! Waiting to unlock ...")

            return result

        return wrapper

    return decorator


@wait_for_file_lock(mode="rb+x", timeout=1200, check_interval=5, fail_when_locked=False)
def safe_write(
    file_path: str,
    sheet_name: str,
    **kwargs,
):
    if "is_new_file" not in kwargs or "locked_file" not in kwargs:
        raise Exception(
            "Locked file instance not provided! Please check if decorator is applied."
        )

    is_new_file = kwargs.pop("is_new_file", False)
    locked_file = kwargs.pop("locked_file")
    interval = kwargs.pop("sleep", random.randint(*SLEEP_INTERVAL))

    if is_new_file:
        excel_output = BytesIO()
        workbook = Workbook()
        excel_kwargs = {"engine": "openpyxl", "mode": "w", **kwargs}
    else:
        excel_output = BytesIO(locked_file.read())
        workbook = load_workbook(excel_output)
        excel_kwargs = {
            "engine": "openpyxl",
            "mode": "a",
            "if_sheet_exists": "overlay",  # "error", "new", "replace", "overlay"
            **kwargs,
        }

    # excel_output.seek(0)
    time.sleep(interval)
    df_data = get_data()

    with pd.ExcelWriter(excel_output, **excel_kwargs) as writer:
        writer.workbook = workbook
        df_data.to_excel(writer, sheet_name=sheet_name, index=False)
        # excel_output.flush()
        # excel_output.seek(0)

    locked_file.seek(0)
    locked_file.truncate(0)
    locked_file.write(excel_output.getvalue())
    locked_file.flush()
    os.fsync(locked_file.fileno())

    return df_data.to_json()


def tested_sequential_code(file_path, sheet_name):
    file_path = Path(file_path)

    if not file_path.exists():
        excel_output = BytesIO()
        workbook = Workbook()
        excel_kwargs = {"engine": "openpyxl", "mode": "w"}
    else:
        with open(file_path, "rb") as f:
            excel_output = BytesIO(f.read())
        workbook = load_workbook(excel_output)
        excel_kwargs = {"engine": "openpyxl", "mode": "a", "if_sheet_exists": "overlay"}

    df = get_data()

    with pd.ExcelWriter(excel_output, **excel_kwargs) as writer:
        print("Wait a minute ...")
        writer.workbook = workbook
        # writer.sheets = {ws.title: ws for ws in workbook.worksheets}
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Save the modified workbook back to the file path
    with open(file_path, "wb") as f:
        f.write(excel_output.getvalue())


def get_data():
    d = {
        "col1": [
            random.randint(*DATA_INTERVAL),
            random.randint(*DATA_INTERVAL),
            random.randint(*DATA_INTERVAL),
        ],
        "col2": [
            random.randint(*DATA_INTERVAL),
            random.randint(*DATA_INTERVAL),
            random.randint(*DATA_INTERVAL),
        ],
    }

    return pd.DataFrame(data=d)


class ThreadWithReturnValue(threading.Thread):
    def __init__(
        self, group=None, target=None, name=None, args=(), kwargs={}, Verbose=None
    ):
        threading.Thread.__init__(self, group, target, name, args, kwargs)
        self.return_value = None

    def run(self):
        if self._target is not None:
            self.return_value = self._target(*self._args, **self._kwargs)


def run_parallel(file_path):
    threads = []

    for i in range(THREADS):
        name = f"Task {i}"
        print(f"STARTING: {name}")
        safe_write.__name__ = name
        thread = ThreadWithReturnValue(
            target=safe_write, args=(file_path, name), name=name
        )
        threads.append(thread)
        thread.start()

    # Wait for all threads to finish
    for i, thread in enumerate(threads):
        name = f"Task {i}"
        print(f"JOINING: {name}")
        thread.join()
        print(f"{name}: {thread.return_value}")


def run_parallel1(file_path):
    with concurrent.futures.ThreadPoolExecutor(max_workers=THREADS) as executor:
        threads = {}

        for i in range(THREADS):
            name = f"Task {i}"
            thread = executor.submit(safe_write, file_path, f"Task {i}")
            threads[thread] = name

        for thread in concurrent.futures.as_completed(threads):
            task = threads[thread]

            try:
                data = thread.result()
            except Exception as exc:
                print(f"{task} generated an exception: {exc}")
            else:
                print(f"{task}: {data}")


def run_sequential(file_path):
    for i in range(THREADS):
        name = f"Task {i}"
        safe_write.__name__ = name
        safe_write(file_path, name, get_data(), sleep=0)


def run_tested_code(file_path):
    for i in range(THREADS):
        name = f"Task {i}"
        tested_sequential_code(file_path, name)


if __name__ == "__main__":
    output_file = "c:\\temp\\test.xlsx"
    # run_tested_code(output_file)
    # run_sequential(output_file)
    run_parallel(output_file)
    # run_parallel1(output_file)
